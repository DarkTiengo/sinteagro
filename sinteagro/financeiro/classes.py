import datetime

from ofxparse import OfxParser
from openpyxl import load_workbook
import xlrd

class Extrato_Reader:

    lines = {}
    caminho = ""

    def open_ofx(self,caminho):
        data = {"Conta": "",
                "Agencia": "",
                "Banco": "",
                "Relatorio": {},
                "Extrato": {}
        }
        id, content, date, type = [],[],[],[]
        with open(caminho) as fileobj:
            ofx = OfxParser.parse(fileobj)
            conta = ofx.account
            data["Conta"] = conta.account_id
            data["Agencia"] = conta.branch_id
            data["Banco"] = conta.institution.organization
            relatorio = conta.statement
            data["Relatorio"].update(inicio = relatorio.start_date, fim = relatorio.end_date),
            for extrato in relatorio.transactions:
                id.append(extrato.id)
                content.append(extrato.memo)
                date.append(extrato.date)
                type.append(extrato.type)
            data["Extrato"].update(id = id,content = content, date = date, type = type)
            return data

    def open_xlsx(self,caminho):
        ws = load_workbook(caminho)
        wb = ws['Relatorio']
        self.common_file = False
        for line in wb:
            self.lines.append(line)
        return True

    def open_xls(self,caminho):
        planilha = xlrd.open_workbook(caminho)
        aba = planilha.sheet_by_index(0)
        type = []
        values = []
        extrato = []
        for r in range(aba.nrows):
            type.append(aba.row_types(r))
            values.append(aba.row_values(r))
        for value in values:
            if type[values.index(value)].count(1) == 0: continue
            try:
                date = datetime.datetime.strptime(value[0],"%d/%m/%Y")
                extrato.append(value)
            except:
                pass
        return extrato

    def open_csv(self,caminho):
        extrato = []
        with open(caminho) as file:
            for row in file.readlines():
                if "Data" in row: continue
                if "S A L D O" in row: continue
                extrato.append(row)
        return extrato

    def open_file(self):
        if ".ofx" in self.caminho:
            self.open_ofx(self.caminho)
        else:
            raise FileExistsError